"""
デバッグ用スクリプト：図形とフォーマットの検証
"""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sys

def debug_shapes(sheet):
    """図形とテキストボックスの内容を確認"""
    print("\n=== 図形・テキストボックスのデバッグ ===")

    # _drawingの確認
    if hasattr(sheet, '_drawing') and sheet._drawing:
        print(f"✓ _drawing が存在します")
        drawing = sheet._drawing
        print(f"  _drawing type: {type(drawing)}")
        print(f"  _drawing attributes: {dir(drawing)}")

        # すべての属性を確認
        for attr in dir(drawing):
            if not attr.startswith('_'):
                try:
                    value = getattr(drawing, attr)
                    if value and not callable(value):
                        print(f"  {attr}: {type(value)}")
                except:
                    pass

        # twoCellAnchorの確認
        if hasattr(drawing, 'twoCellAnchor'):
            print(f"\n  twoCellAnchor の数: {len(drawing.twoCellAnchor) if drawing.twoCellAnchor else 0}")

            for idx, anchor in enumerate(drawing.twoCellAnchor):
                print(f"\n  === Anchor {idx + 1} ===")
                print(f"    type: {type(anchor)}")
                print(f"    attributes: {[a for a in dir(anchor) if not a.startswith('_')]}")

                # spの確認
                if hasattr(anchor, 'sp') and anchor.sp:
                    sp = anchor.sp
                    print(f"    ✓ sp (shape) が存在")
                    print(f"      sp attributes: {[a for a in dir(sp) if not a.startswith('_')]}")

                    # 名前の取得
                    if hasattr(sp, 'nvSpPr') and sp.nvSpPr:
                        if hasattr(sp.nvSpPr, 'cNvPr') and sp.nvSpPr.cNvPr:
                            name = getattr(sp.nvSpPr.cNvPr, 'name', 'No name')
                            print(f"      名前: {name}")

                    # テキストの取得
                    if hasattr(sp, 'txBody') and sp.txBody:
                        print(f"      ✓ txBody が存在")
                        txBody = sp.txBody
                        print(f"        txBody attributes: {[a for a in dir(txBody) if not a.startswith('_')]}")

                        if hasattr(txBody, 'p'):
                            print(f"        paragraphs の数: {len(txBody.p) if txBody.p else 0}")
                            for p_idx, paragraph in enumerate(txBody.p):
                                print(f"        Paragraph {p_idx + 1}:")
                                if hasattr(paragraph, 'r'):
                                    for r_idx, run in enumerate(paragraph.r):
                                        if hasattr(run, 't') and run.t:
                                            print(f"          Run {r_idx + 1} text: '{run.t}'")

        # oneCellAnchorの確認
        if hasattr(drawing, 'oneCellAnchor'):
            print(f"\n  oneCellAnchor の数: {len(drawing.oneCellAnchor) if drawing.oneCellAnchor else 0}")

        # absoluteAnchorの確認
        if hasattr(drawing, 'absoluteAnchor'):
            print(f"\n  absoluteAnchor の数: {len(drawing.absoluteAnchor) if drawing.absoluteAnchor else 0}")
    else:
        print("✗ _drawing が存在しません")


def debug_cell_format(sheet, sheet_with_values, cell_address):
    """特定のセルのフォーマット情報を確認"""
    print(f"\n=== セル {cell_address} のフォーマットデバッグ ===")

    # 数式用ワークブックから取得
    cell = sheet[cell_address]
    print(f"数式用セル:")
    print(f"  値: {cell.value}")
    print(f"  データ型: {cell.data_type}")
    print(f"  number_format: {cell.number_format}")
    print(f"  is_date (属性): {hasattr(cell, 'is_date')}")
    if hasattr(cell, 'is_date'):
        print(f"  is_date: {cell.is_date}")

    # 実数値用ワークブックから取得
    if sheet_with_values:
        value_cell = sheet_with_values[cell_address]
        print(f"\n実数値用セル:")
        print(f"  値: {value_cell.value}")
        print(f"  値の型: {type(value_cell.value)}")
        print(f"  データ型: {value_cell.data_type}")
        print(f"  number_format: {value_cell.number_format}")
        print(f"  is_date (属性): {hasattr(value_cell, 'is_date')}")
        if hasattr(value_cell, 'is_date'):
            print(f"  is_date: {value_cell.is_date}")

        # 数値の場合、日付変換を試みる
        if isinstance(value_cell.value, (int, float)):
            try:
                value = value_cell.value
                if value > 59:
                    excel_date = datetime(1899, 12, 30) + timedelta(days=value)
                elif value >= 1:
                    excel_date = datetime(1899, 12, 31) + timedelta(days=value)
                else:
                    excel_date = datetime(1900, 1, 1) + timedelta(days=value)
                print(f"\n  日付変換テスト: {excel_date.strftime('%Y年%m月%d日')}")
            except Exception as e:
                print(f"\n  日付変換エラー: {e}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python test_debug.py <Excelファイルパス> [セルアドレス]")
        sys.exit(1)

    excel_path = sys.argv[1]
    cell_address = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"Excelファイルを読み込み中: {excel_path}")

    # ワークブックを読み込み
    wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
    wb_values = openpyxl.load_workbook(excel_path, data_only=True)

    # 最初のシートを取得
    sheet = wb_formulas.active
    sheet_with_values = wb_values.active

    print(f"シート名: {sheet.title}")

    # 図形のデバッグ
    debug_shapes(sheet)

    # セルのフォーマットデバッグ
    if cell_address:
        debug_cell_format(sheet, sheet_with_values, cell_address)

    print("\n=== デバッグ完了 ===")
