"""
Table Parser - 表検出・変換ロジック
"""

from typing import List, Tuple, Dict, Any
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import re


class TableParser:
    """テーブル解析・変換クラス"""

    def __init__(self, config: Dict[str, Any]):
        self.config = config

    def parse_tables(self, sheet, sheet_with_values=None) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        シート内のテーブルを検出して変換

        Args:
            sheet: openpyxlのWorksheetオブジェクト（数式情報用）
            sheet_with_values: 実数値を含むシート（data_only=True）

        Returns:
            (Markdown形式のテーブルリスト, テーブル情報のリスト)
        """
        tables_md = []
        tables_info = []

        # Excelの正式なテーブルオブジェクトがあればそれを使用
        if hasattr(sheet, 'tables') and sheet.tables:
            for table_name, table_range in sheet.tables.items():
                md_table = self.convert_range_to_markdown(sheet, table_range, sheet_with_values)
                if md_table:
                    tables_md.append(md_table)
                    tables_info.append({
                        'name': table_name,
                        'range': table_range,
                        'type': 'excel_table',
                        'markdown': md_table  # Phase 3: AI機能用にMarkdownを保存
                    })
        else:
            # テーブルオブジェクトがない場合は、空白行で区切られた複数の表を検出
            if sheet.dimensions and sheet.dimensions != "A1:A1":
                detected_tables = self._detect_tables_by_blank_rows(sheet, sheet_with_values)
                for idx, table_range in enumerate(detected_tables):
                    md_table = self.convert_range_to_markdown(sheet, table_range, sheet_with_values)
                    if md_table:
                        tables_md.append(md_table)
                        tables_info.append({
                            'name': f'table_{idx + 1}',
                            'range': table_range,
                            'type': 'auto_detected',
                            'markdown': md_table  # Phase 3: AI機能用にMarkdownを保存
                        })

        return tables_md, tables_info

    def _detect_tables_by_blank_rows(self, sheet, sheet_with_values=None) -> List[str]:
        """
        空白行で区切られた複数の表を検出

        Args:
            sheet: openpyxlのWorksheetオブジェクト
            sheet_with_values: 実数値を含むシート

        Returns:
            検出されたテーブル範囲のリスト（例: ["A1:D5", "A7:D10"]）
        """
        if not sheet.dimensions or sheet.dimensions == "A1:A1":
            return []

        # 使用範囲を取得
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column

        # 空白行を検出
        blank_rows = []
        for row_idx in range(min_row, max_row + 1):
            is_blank = True
            for col_idx in range(min_col, max_col + 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    is_blank = False
                    break
            if is_blank:
                blank_rows.append(row_idx)

        # 空白行がない場合は、全体を1つのテーブルとして扱う
        if not blank_rows:
            return [sheet.dimensions]

        # 空白行で区切られたテーブル範囲を生成
        table_ranges = []
        current_start_row = min_row

        for blank_row in blank_rows:
            # 空白行の前までをテーブルとして扱う
            if blank_row > current_start_row:
                # テーブル範囲を作成
                start_cell = f"{get_column_letter(min_col)}{current_start_row}"
                end_cell = f"{get_column_letter(max_col)}{blank_row - 1}"
                table_range = f"{start_cell}:{end_cell}"
                table_ranges.append(table_range)

            # 次のテーブルの開始行を設定
            current_start_row = blank_row + 1

        # 最後のテーブルを追加
        if current_start_row <= max_row:
            start_cell = f"{get_column_letter(min_col)}{current_start_row}"
            end_cell = f"{get_column_letter(max_col)}{max_row}"
            table_range = f"{start_cell}:{end_cell}"
            table_ranges.append(table_range)

        # 空のテーブルを除外
        valid_ranges = []
        for table_range in table_ranges:
            # テーブル範囲が空でないか確認
            has_data = False
            for row in sheet[table_range]:
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        has_data = True
                        break
                if has_data:
                    break
            if has_data:
                valid_ranges.append(table_range)

        return valid_ranges if valid_ranges else [sheet.dimensions]

    def convert_range_to_markdown(self, sheet, cell_range: str, sheet_with_values=None) -> str:
        """
        セル範囲をMarkdownテーブルに変換

        Args:
            sheet: openpyxlのWorksheetオブジェクト（数式情報用）
            cell_range: セル範囲（例: "A1:D10"）
            sheet_with_values: 実数値を含むシート（data_only=True）

        Returns:
            Markdown形式のテーブル
        """
        try:
            # セル範囲からデータを取得
            data = []
            formulas = {}  # 数式を保存 {(row, col): formula}

            for row_idx, row in enumerate(sheet[cell_range]):
                row_data = []
                for col_idx, cell in enumerate(row):
                    # 実数値の取得を優先
                    value = None
                    format_cell = cell  # フォーマット情報を持つセル

                    if sheet_with_values:
                        # 実数値シートから値を取得
                        value_cell = sheet_with_values.cell(cell.row, cell.column)
                        value = value_cell.value
                    else:
                        value = cell.value

                    # 数式の確認と保存
                    if cell.data_type == 'f':  # 数式セル
                        formula = str(cell.value)
                        formulas[(row_idx, col_idx)] = {
                            'cell': cell.coordinate,
                            'formula': formula
                        }
                        # 実数値がNoneの場合は数式を表示
                        if value is None:
                            value = formula

                    # セルのフォーマットを適用して値を整形
                    # フォーマット情報は元のセル（数式セル）から取得
                    if value is not None and value != "":
                        value = self._format_cell_value(value, format_cell)

                    if value is None:
                        value = ""
                    # セル内の改行を<br>タグに置換してMarkdownテーブルの構造を保持
                    cell_text = str(value).replace('\n', '<br>').replace('\r', '')
                    row_data.append(cell_text)
                data.append(row_data)

            if not data:
                return ""

            # pandasのDataFrameに変換
            df = pd.DataFrame(data)

            # 空の行と列を削除
            # NAを使うとdf.emptyでエラーが出るため、Noneを使用
            df = df.replace('', None)
            df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)

            # emptyの代わりにshapeでチェック
            if df.shape[0] == 0 or df.shape[1] == 0:
                return ""

            # NAをfillnaで空文字に戻す
            df = df.fillna('')

            # ヘッダーの検出
            if self.config.get('detect_header', True) and len(df) > 0:
                # 最初の行をヘッダーとして使用
                headers = df.iloc[0].tolist()
                df = df.iloc[1:]
                df.columns = headers
            else:
                # ヘッダーなし（列番号を使用）
                df.columns = [f"Column {i+1}" for i in range(len(df.columns))]

            # Markdown形式に変換
            md_table = df.to_markdown(index=False, tablefmt='github')

            # 数式の備考を追加（設定により）
            if formulas and self.config.get('show_formulas', True):
                formula_notes = self._generate_formula_notes(formulas)
                if formula_notes:
                    md_table += f"\n\n{formula_notes}"

            # 要約の生成（設定により）
            if self.config.get('generate_summary', False):
                summary = self._generate_table_summary(df)
                if summary:
                    md_table = f"{summary}\n\n{md_table}"

            return md_table

        except Exception as e:
            import traceback
            error_msg = f"テーブル変換エラー: {str(e)}"
            print(error_msg)

            # デバッグモードの場合は詳細なエラー情報を出力
            if self.config.get('verbose_logging', False):
                print(f"セル範囲: {cell_range}")
                print(f"エラー詳細:")
                traceback.print_exc()

            # エラーが発生した場合は空文字列を返して処理を続行
            return ""

    def _generate_formula_notes(self, formulas: Dict) -> str:
        """数式の備考を生成"""
        if not formulas:
            return ""

        notes = ["**【数式備考】**"]
        for (row, col), info in sorted(formulas.items()):
            cell_ref = info['cell']
            formula = info['formula']
            notes.append(f"- セル {cell_ref}: `{formula}`")

        return '\n'.join(notes)

    def _generate_table_summary(self, df: pd.DataFrame) -> str:
        """テーブルの要約を生成"""
        summary_parts = []

        # 基本情報
        summary_parts.append(f"データ行数: {len(df)}行")

        # 数値列の情報
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if numeric_cols:
            summary_parts.append(f"数値列: {', '.join(numeric_cols)}")

        if summary_parts:
            return "【テーブル要約】 " + "、".join(summary_parts)

        return ""

    def _format_cell_value(self, value, cell):
        """
        セルの値をフォーマットして返す

        Args:
            value: セルの値
            cell: openpyxlのCellオブジェクト

        Returns:
            フォーマットされた値
        """
        # datetimeオブジェクトの場合はそのまま文字列化
        if isinstance(value, datetime):
            # セルのフォーマットを確認
            if cell.number_format:
                format_str = cell.number_format
                # 日付のみのフォーマット
                if any(x in format_str.lower() for x in ['yy', 'mm', 'dd']) and 'h' not in format_str.lower():
                    return value.strftime('%Y年%m月%d日')
                # 時刻を含むフォーマット
                elif 'h' in format_str.lower() or 'm' in format_str.lower() or 's' in format_str.lower():
                    return value.strftime('%Y年%m月%d日 %H:%M:%S')
            # デフォルトのフォーマット
            return value.strftime('%Y年%m月%d日')

        # 数値の場合、日付フォーマットかどうかを確認
        if isinstance(value, (int, float)):
            # 日付フォーマットかどうかを判定
            is_date_format = self._is_date_format(cell, value)

            if is_date_format:
                try:
                    # Excelの日付シリアル値を変換（1900年1月1日からの経過日数）
                    excel_date = self._convert_excel_date(value)

                    # 時刻を含むかチェック
                    format_str = cell.number_format if cell.number_format else ''
                    has_time = 'h' in format_str.lower() or ('m' in format_str.lower() and ':' in format_str) or \
                              '時' in format_str or '分' in format_str or '秒' in format_str

                    if has_time:
                        return excel_date.strftime('%Y年%m月%d日 %H:%M:%S')
                    else:
                        return excel_date.strftime('%Y年%m月%d日')
                except (ValueError, OverflowError) as e:
                    # 変換に失敗した場合は元の値を返す
                    print(f"日付変換エラー（値: {value}）: {str(e)}")
                    pass

            # 日付でない場合、他のフォーマットをチェック
            if cell.number_format:
                format_str = cell.number_format

                # パーセンテージフォーマットの検出
                if '%' in format_str:
                    try:
                        return f"{value * 100:.2f}%"
                    except:
                        pass

                # 通貨フォーマットの検出
                if '¥' in format_str or '円' in format_str:
                    try:
                        return f"¥{value:,.0f}"
                    except:
                        pass

                # カンマ区切りの数値フォーマット
                if '#,##0' in format_str or '0.00' in format_str:
                    try:
                        if '.' in format_str:
                            decimal_places = format_str.count('0') - format_str.index('.') - 1
                            return f"{value:,.{decimal_places}f}"
                        else:
                            return f"{value:,.0f}"
                    except:
                        pass

        return value

    def _is_date_format(self, cell, value) -> bool:
        """
        セルが日付フォーマットかどうかを判定

        Args:
            cell: openpyxlのCellオブジェクト
            value: セルの値

        Returns:
            日付フォーマットの場合True
        """
        # 数値の範囲チェック（Excelの日付シリアル値の妥当な範囲）
        # 1（1900/1/1）から73050（2100/1/1）程度
        if not isinstance(value, (int, float)) or value < 0 or value > 73050:
            return False

        # 1. ビルトイン日付フォーマットIDをチェック
        # Excelの日付/時刻フォーマットのビルトインID: 14-22, 27-36, 45-47, 50-58
        try:
            if hasattr(cell, '_style'):
                numFmtId = cell._style.numFmtId
                # 日付/時刻のビルトインフォーマットID
                date_format_ids = list(range(14, 23)) + list(range(27, 37)) + list(range(45, 48)) + list(range(50, 59))
                if numFmtId in date_format_ids:
                    return True
        except:
            pass

        # 2. number_formatの文字列をチェック
        if cell.number_format:
            format_str = cell.number_format

            # 日付関連のキーワードを含むかチェック
            # 年月日の英字フォーマット
            has_date_keywords = any(x in format_str.lower() for x in ['yy', 'mm', 'dd', 'yyyy', 'mmmm', 'mmmmm'])
            # 日本語のフォーマット
            has_japanese_date = any(x in format_str for x in ['年', '月', '日'])

            if has_date_keywords or has_japanese_date:
                # ただし、時刻のみのフォーマット（h:mm, m:ss等）は除外
                # "m"が":"の前後にある場合は分（minute）、それ以外は月（month）
                if ':' in format_str and 'm' in format_str.lower():
                    # h:mm や m:ss のような時刻フォーマット
                    # 日付も含まれているかチェック
                    if has_japanese_date or 'yy' in format_str.lower() or 'dd' in format_str.lower():
                        return True
                    # 時刻のみの場合は日付フォーマットではない
                    return False
                else:
                    return True

        # 3. openpyxlのis_date属性をチェック（最後の手段）
        try:
            if hasattr(cell, 'is_date') and cell.is_date:
                return True
        except:
            pass

        return False

    def _convert_excel_date(self, value: float) -> datetime:
        """
        Excelの日付シリアル値をdatetimeオブジェクトに変換

        Args:
            value: Excelの日付シリアル値

        Returns:
            datetimeオブジェクト
        """
        # Excelは1900年1月1日を1とする（ただし1900年をうるう年として誤って扱う）
        # 1900年2月28日までは正しいが、1900年2月29日（存在しない）をカウントしている
        # そのため、1900年3月1日（シリアル値60）以降は調整が必要

        if value > 59:
            # 1900年3月1日以降（Excel のバグ考慮後）
            # 基準日: 1899年12月30日（シリアル値0に相当）
            return datetime(1899, 12, 30) + timedelta(days=value)
        elif value >= 1:
            # 1900年1月1日から2月28日まで
            # 基準日: 1899年12月31日
            return datetime(1899, 12, 31) + timedelta(days=value)
        else:
            # 1未満は時刻のみ（日付部分が1900年1月1日）
            return datetime(1900, 1, 1) + timedelta(days=value)
