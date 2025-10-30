"""
Excel Sheet to Image Converter
シートをPDF/画像形式に変換するモジュール
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.table import Table
from matplotlib import font_manager
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
import numpy as np
from typing import List, Dict, Optional, Tuple
import os
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from io import BytesIO


def setup_japanese_font():
    """日本語フォントの設定"""
    # 日本語フォントの候補リスト
    japanese_fonts = [
        'Noto Sans CJK JP',
        'Noto Sans JP',
        'IPAexGothic',
        'IPAGothic',
        'Hiragino Sans',
        'Yu Gothic',
        'Meiryo',
        'MS Gothic',
        'TakaoPGothic',
        'VL Gothic',
        'DejaVu Sans'
    ]

    # 利用可能なフォントを取得
    available_fonts = [f.name for f in font_manager.fontManager.ttflist]

    # 日本語フォントを探す
    for font in japanese_fonts:
        if font in available_fonts:
            matplotlib.rcParams['font.family'] = font
            return font

    # フォントが見つからない場合はsans-serifのフォールバック
    matplotlib.rcParams['font.family'] = 'sans-serif'
    matplotlib.rcParams['font.sans-serif'] = japanese_fonts + ['DejaVu Sans']
    # マイナス記号の文字化け対策
    matplotlib.rcParams['axes.unicode_minus'] = False

    return 'sans-serif'


class SheetToImageConverter:
    """ExcelシートをPDFや画像に変換するクラス"""

    def __init__(self, dpi: int = 150, page_size: str = "A4"):
        """
        初期化

        Args:
            dpi: 画像の解像度（デフォルト: 150）
            page_size: ページサイズ（デフォルト: "A4"）
        """
        self.dpi = dpi
        self.page_size = page_size

        # 日本語フォントの設定
        self.font_name = setup_japanese_font()

    def excel_to_images(self, excel_path: str, output_dir: str) -> Dict[str, List[str]]:
        """
        Excelファイルの各シートを画像に変換

        Args:
            excel_path: Excelファイルのパス
            output_dir: 出力ディレクトリ

        Returns:
            シート名をキーとした画像ファイルパスのリスト
        """
        os.makedirs(output_dir, exist_ok=True)

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_images = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # シートを画像化
            image_path = self._sheet_to_image(sheet, output_dir, sheet_name)
            sheet_images[sheet_name] = [image_path]

        workbook.close()
        return sheet_images

    def _sheet_to_image(self, sheet, output_dir: str, sheet_name: str) -> str:
        """
        シートを1つの画像に変換

        Args:
            sheet: openpyxlのWorksheetオブジェクト
            output_dir: 出力ディレクトリ
            sheet_name: シート名

        Returns:
            画像ファイルのパス
        """
        # データ範囲を取得
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            # 空のシートの場合
            return self._create_empty_sheet_image(output_dir, sheet_name)

        # セルデータを取得
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            data.append(row_data)

        # 画像として描画
        image_path = os.path.join(output_dir, f"{self._sanitize_filename(sheet_name)}.png")
        self._render_table_as_image(data, image_path, sheet_name)

        return image_path

    def _render_table_as_image(self, data: List[List[str]], output_path: str, title: str):
        """
        テーブルデータを画像として描画

        Args:
            data: 2次元配列のテーブルデータ
            output_path: 出力画像パス
            title: タイトル
        """
        if not data or not data[0]:
            self._create_empty_sheet_image(os.path.dirname(output_path), title)
            return

        rows = len(data)
        cols = len(data[0])

        # 図のサイズを計算（セル数に応じて調整）
        cell_width = 2.0
        cell_height = 0.5
        fig_width = min(max(cols * cell_width, 8), 20)
        fig_height = min(max(rows * cell_height + 1, 6), 30)

        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=self.dpi)
        ax.axis('tight')
        ax.axis('off')

        # タイトルを追加
        ax.set_title(title, fontsize=14, fontweight='bold', pad=20)

        # テーブルを作成
        # データの各セルの長さを制限
        display_data = []
        for row in data:
            display_row = []
            for cell in row:
                cell_str = str(cell)
                if len(cell_str) > 50:
                    cell_str = cell_str[:47] + "..."
                display_row.append(cell_str)
            display_data.append(display_row)

        table = ax.table(
            cellText=display_data,
            cellLoc='left',
            loc='center',
            bbox=[0, 0, 1, 1]
        )

        # テーブルのスタイル設定
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1, 1.5)

        # ヘッダー行のスタイル
        for i in range(cols):
            cell = table[(0, i)]
            cell.set_facecolor('#4CAF50')
            cell.set_text_props(weight='bold', color='white')

        # 交互に背景色を設定
        for i in range(1, rows):
            for j in range(cols):
                cell = table[(i, j)]
                if i % 2 == 0:
                    cell.set_facecolor('#f0f0f0')
                else:
                    cell.set_facecolor('white')

        plt.tight_layout()
        plt.savefig(output_path, dpi=self.dpi, bbox_inches='tight', facecolor='white')
        plt.close()

    def _create_empty_sheet_image(self, output_dir: str, sheet_name: str) -> str:
        """
        空のシート用の画像を作成

        Args:
            output_dir: 出力ディレクトリ
            sheet_name: シート名

        Returns:
            画像ファイルのパス
        """
        image_path = os.path.join(output_dir, f"{self._sanitize_filename(sheet_name)}.png")

        # 空白の画像を作成
        img = Image.new('RGB', (800, 600), color='white')
        draw = ImageDraw.Draw(img)

        # 日本語対応フォントを試す
        font_paths = [
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf",
            "/usr/share/fonts/truetype/takao-gothic/TakaoPGothic.ttf",
            "/usr/share/fonts/truetype/vlgothic/VL-Gothic-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
            "C:\\Windows\\Fonts\\msgothic.ttc",
            "C:\\Windows\\Fonts\\YuGothR.ttc"
        ]

        font = None
        for font_path in font_paths:
            try:
                font = ImageFont.truetype(font_path, 24)
                break
            except:
                continue

        if font is None:
            try:
                font = ImageFont.load_default()
            except:
                font = ImageFont.load_default()

        text = f"空のシート: {sheet_name}"
        # テキストの位置を計算（中央）
        try:
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
        except:
            # フォールバック
            text_width = len(text) * 12
            text_height = 24

        x = (800 - text_width) / 2
        y = (600 - text_height) / 2

        draw.text((x, y), text, fill='black', font=font)

        img.save(image_path)
        return image_path

    def images_to_pdf(self, image_paths: List[str], output_pdf: str):
        """
        複数の画像を1つのPDFに結合

        Args:
            image_paths: 画像ファイルパスのリスト
            output_pdf: 出力PDFパス
        """
        from PIL import Image

        if not image_paths:
            return

        images = []
        for img_path in image_paths:
            img = Image.open(img_path)
            # RGBモードに変換（PDFに変換するため）
            if img.mode != 'RGB':
                img = img.convert('RGB')
            images.append(img)

        if images:
            # 最初の画像を使ってPDFを作成し、残りを追加
            images[0].save(output_pdf, save_all=True, append_images=images[1:])

    def sheet_to_pdf_via_reportlab(self, sheet, output_pdf: str, sheet_name: str):
        """
        ReportLabを使ってシートをPDFに直接変換（より詳細な制御）

        Args:
            sheet: openpyxlのWorksheetオブジェクト
            output_pdf: 出力PDFパス
            sheet_name: シート名
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        # データ取得
        max_row = sheet.max_row
        max_col = sheet.max_column

        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            data.append(row_data)

        # PDF作成
        doc = SimpleDocTemplate(output_pdf, pagesize=landscape(A4))
        elements = []

        # タイトル
        styles = getSampleStyleSheet()
        title = Paragraph(f"<b>{sheet_name}</b>", styles['Title'])
        elements.append(title)

        # テーブル
        if data:
            t = Table(data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t)

        doc.build(elements)

    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        """
        ファイル名として使えない文字を置換

        Args:
            filename: 元のファイル名

        Returns:
            サニタイズされたファイル名
        """
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename
