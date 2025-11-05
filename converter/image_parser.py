"""
Image Parser - 画像・グラフ・図形抽出ロジック
"""

import os
from typing import List, Tuple, Dict, Any
from PIL import Image
import io
import zipfile
from xml.etree import ElementTree as ET
import re


class ImageParser:
    """画像・グラフ・図形抽出クラス"""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.image_counter = 0
        self.shape_counter = 0
        self.excel_file_path = None  # Excelファイルのパスを保持

    def extract_images(self, sheet) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        シートから画像・グラフを抽出

        Args:
            sheet: openpyxlのWorksheetオブジェクト

        Returns:
            (Markdown形式の画像参照リスト, 画像情報のリスト)
        """
        images_md = []
        images_info = []

        if not hasattr(sheet, '_images') or not sheet._images:
            return images_md, images_info

        output_dir = self.config.get('output_dir', 'images')

        # 出力ディレクトリの作成
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        for img_idx, img in enumerate(sheet._images):
            try:
                self.image_counter += 1

                # 画像ファイル名の生成
                image_format = self.config.get('image_format', 'png')
                image_filename = f"chart_{self.image_counter:03d}.{image_format}"
                image_path = os.path.join(output_dir, image_filename)

                # 画像を保存
                self._save_image(img, image_path)

                # Markdown形式の画像参照を生成
                title = getattr(img, 'name', None) or f"Image {self.image_counter}"
                md_image = f"![{title}](./{image_path})"

                # 画像説明の生成（設定により）
                if self.config.get('generate_image_description', False):
                    description = self._generate_image_description(img)
                    if description:
                        md_image += f"\n\n{description}"

                images_md.append(md_image)
                images_info.append({
                    'index': self.image_counter,
                    'filename': image_filename,
                    'path': image_path,
                    'title': title,
                    'type': 'image'
                })

            except Exception as e:
                print(f"画像抽出エラー: {str(e)}")
                continue

        return images_md, images_info

    def extract_shapes(self, sheet, excel_path: str = None) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        シートから図形とその中のテキストを抽出（テキストボックスを含む）

        Args:
            sheet: openpyxlのWorksheetオブジェクト
            excel_path: Excelファイルのパス（ZIPベースの抽出に使用）

        Returns:
            (Markdown形式の図形情報リスト, 図形情報のリスト)
        """
        shapes_md = []
        shapes_info = []

        # 方法1: openpyxlの_drawingを使用
        openpyxl_shapes_md, openpyxl_shapes_info = self._extract_shapes_from_openpyxl(sheet)

        # 方法2: openpyxlで取得できなかった場合、ZIPベースで抽出
        if not openpyxl_shapes_info and excel_path:
            if self.config.get('verbose_logging', False):
                print(f"openpyxlで図形が取得できなかったため、ZIP解析を試行します...")

            zip_shapes_md, zip_shapes_info = self._extract_shapes_from_zip(excel_path, sheet.title)
            shapes_md.extend(zip_shapes_md)
            shapes_info.extend(zip_shapes_info)
        else:
            shapes_md.extend(openpyxl_shapes_md)
            shapes_info.extend(openpyxl_shapes_info)

        # 抽出結果をログ出力
        if shapes_info:
            print(f"✓ {len(shapes_info)}個の図形を抽出しました")
            if self.config.get('verbose_logging', False):
                for shape in shapes_info:
                    print(f"  - {shape['name']}: {len(shape.get('text', ''))}文字のテキスト")

        return shapes_md, shapes_info

    def _extract_shapes_from_openpyxl(self, sheet) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        openpyxlの_drawingを使用して図形を抽出

        Args:
            sheet: openpyxlのWorksheetオブジェクト

        Returns:
            (Markdown形式の図形情報リスト, 図形情報のリスト)
        """
        shapes_md = []
        shapes_info = []

        # openpyxlの図形オブジェクトにアクセス（_drawingを使用）
        if not hasattr(sheet, '_drawing') or not sheet._drawing:
            return shapes_md, shapes_info

        try:
            drawing = sheet._drawing

            # すべてのアンカータイプをチェック（テキストボックスはどのアンカータイプでも存在する可能性がある）
            anchor_lists = []

            if hasattr(drawing, 'twoCellAnchor') and drawing.twoCellAnchor:
                anchor_lists.append(('twoCellAnchor', drawing.twoCellAnchor))

            if hasattr(drawing, 'oneCellAnchor') and drawing.oneCellAnchor:
                anchor_lists.append(('oneCellAnchor', drawing.oneCellAnchor))

            if hasattr(drawing, 'absoluteAnchor') and drawing.absoluteAnchor:
                anchor_lists.append(('absoluteAnchor', drawing.absoluteAnchor))

            # すべてのアンカーから図形を抽出
            for anchor_type, anchors in anchor_lists:
                for anchor in anchors:
                    try:
                        anchor_info = self._get_anchor_info(anchor)
                        shapes_to_process = []

                        # 方法1: 単一の図形（sp）を取得
                        if hasattr(anchor, 'sp') and anchor.sp:
                            shapes_to_process.append((anchor.sp, False))

                        # 方法2: グループ化された図形（grpSp）を取得
                        if hasattr(anchor, 'grpSp') and anchor.grpSp:
                            # グループ内のすべての図形を取得
                            if hasattr(anchor.grpSp, 'sp'):
                                group_shapes = anchor.grpSp.sp if isinstance(anchor.grpSp.sp, list) else [anchor.grpSp.sp]
                                for grp_shape in group_shapes:
                                    if grp_shape:
                                        shapes_to_process.append((grp_shape, True))

                        # すべての図形を処理
                        for shape, is_grouped in shapes_to_process:
                            if not shape:
                                continue

                            self.shape_counter += 1

                            # 図形の基本情報
                            shape_data = {
                                'index': self.shape_counter,
                                'type': 'shape',
                                'anchor_type': anchor_type,
                                'is_grouped': is_grouped
                            }

                            # 図形名を取得
                            shape_name = f"Shape {self.shape_counter}"
                            if hasattr(shape, 'nvSpPr') and shape.nvSpPr:
                                if hasattr(shape.nvSpPr, 'cNvPr') and shape.nvSpPr.cNvPr:
                                    name = getattr(shape.nvSpPr.cNvPr, 'name', None)
                                    if name:
                                        shape_name = name

                            shape_data['name'] = shape_name

                            # 図形内のテキストを取得
                            shape_text = self._extract_text_from_shape(shape)

                            # テキストが取得できた場合のみMarkdownに追加
                            if shape_text:
                                shape_data['text'] = shape_text

                                # Markdown形式で出力
                                group_indicator = " (グループ化)" if is_grouped else ""
                                md_parts = [f"### 📐 {shape_name}{group_indicator}"]
                                # テキストを引用として表示（複数行対応）
                                for line in shape_text.split('\n'):
                                    if line.strip():
                                        md_parts.append(f"> {line}")

                                # 位置情報を追加
                                if anchor_info:
                                    shape_data['position'] = anchor_info
                                    md_parts.append(f"\n**位置情報**: {anchor_info}")

                                md_shape = '\n'.join(md_parts)
                                shapes_md.append(md_shape)
                                shapes_info.append(shape_data)

                    except Exception as e:
                        print(f"図形抽出エラー（{anchor_type}）: {str(e)}")
                        import traceback
                        if self.config.get('verbose_logging', False):
                            traceback.print_exc()
                        continue

        except Exception as e:
            print(f"図形抽出全体エラー: {str(e)}")
            import traceback
            if self.config.get('verbose_logging', False):
                traceback.print_exc()

        # 抽出結果をログ出力
        if shapes_info:
            print(f"✓ {len(shapes_info)}個の図形を抽出しました")
            if self.config.get('verbose_logging', False):
                for shape in shapes_info:
                    print(f"  - {shape['name']}: {len(shape.get('text', ''))}文字のテキスト")

        return shapes_md, shapes_info

    def _extract_text_from_shape(self, shape) -> str:
        """
        図形からテキストを抽出

        Args:
            shape: 図形オブジェクト

        Returns:
            抽出されたテキスト
        """
        text_parts = []

        try:
            # txBodyからテキストを取得
            if hasattr(shape, 'txBody') and shape.txBody:
                txBody = shape.txBody

                # 段落（paragraph）のリストを取得
                paragraphs = []
                if hasattr(txBody, 'p'):
                    if isinstance(txBody.p, list):
                        paragraphs = txBody.p
                    else:
                        paragraphs = [txBody.p]

                # 各段落からテキストを抽出
                for paragraph in paragraphs:
                    if paragraph is None:
                        continue

                    paragraph_text = []

                    # run（テキストの塊）のリストを取得
                    runs = []
                    if hasattr(paragraph, 'r'):
                        if isinstance(paragraph.r, list):
                            runs = paragraph.r
                        else:
                            runs = [paragraph.r]

                    # 各runからテキストを抽出
                    for run in runs:
                        if run is None:
                            continue

                        if hasattr(run, 't') and run.t:
                            paragraph_text.append(str(run.t))

                    # 段落のテキストを結合
                    if paragraph_text:
                        text_parts.append(''.join(paragraph_text))

        except Exception as e:
            print(f"テキスト抽出エラー: {str(e)}")
            if self.config.get('verbose_logging', False):
                import traceback
                traceback.print_exc()

        # 段落を改行で結合
        return '\n'.join(text_parts) if text_parts else None

    def _extract_shapes_from_zip(self, excel_path: str, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        ZIPファイルとしてExcelを開き、XMLから直接図形を抽出

        Args:
            excel_path: Excelファイルのパス
            sheet_name: シート名

        Returns:
            (Markdown形式の図形情報リスト, 図形情報のリスト)
        """
        shapes_md = []
        shapes_info = []

        if not excel_path or not os.path.exists(excel_path):
            return shapes_md, shapes_info

        try:
            with zipfile.ZipFile(excel_path, 'r') as zip_ref:
                # シートインデックスを取得するため、workbook.xmlを読む
                workbook_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
                wb_root = ET.fromstring(workbook_xml)

                # 名前空間の定義
                ns = {
                    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }

                # シート名からシートインデックスを取得
                sheet_index = None
                sheets = wb_root.findall('.//main:sheet', ns)
                for idx, sheet_elem in enumerate(sheets, 1):
                    name = sheet_elem.get('name')
                    if name == sheet_name:
                        sheet_index = idx
                        break

                if sheet_index is None:
                    return shapes_md, shapes_info

                # drawing*.xmlファイルを探す
                drawing_files = [f for f in zip_ref.namelist()
                                if f.startswith('xl/drawings/drawing') and f.endswith('.xml')]

                if not drawing_files:
                    return shapes_md, shapes_info

                # シートに対応するdrawingファイルを見つける
                # worksheet*.xml.relsを確認
                rels_path = f'xl/worksheets/_rels/sheet{sheet_index}.xml.rels'
                drawing_rel_id = None

                try:
                    rels_content = zip_ref.read(rels_path).decode('utf-8')
                    rels_root = ET.fromstring(rels_content)
                    rels_ns = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}

                    for rel in rels_root.findall('.//rel:Relationship', rels_ns):
                        if 'drawing' in rel.get('Type', '').lower():
                            target = rel.get('Target')
                            # ../drawings/drawing1.xml のような形式
                            drawing_file = 'xl/drawings/' + target.split('/')[-1]
                            if drawing_file in zip_ref.namelist():
                                drawing_rel_id = drawing_file
                                break
                except:
                    # relsファイルがない場合は、最初のdrawingファイルを使用
                    if drawing_files:
                        drawing_rel_id = drawing_files[0]

                if not drawing_rel_id:
                    return shapes_md, shapes_info

                # drawingファイルを解析
                drawing_content = zip_ref.read(drawing_rel_id).decode('utf-8')
                drawing_root = ET.fromstring(drawing_content)

                # 名前空間の定義
                drawing_ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                }

                # すべてのアンカータイプから図形を探す
                anchor_types = ['twoCellAnchor', 'oneCellAnchor', 'absoluteAnchor']

                for anchor_type in anchor_types:
                    anchors = drawing_root.findall(f'.//xdr:{anchor_type}', drawing_ns)

                    for anchor in anchors:
                        # アンカーの位置情報を取得（グループ内の図形にも使用）
                        anchor_position = self._get_position_from_xml_anchor(anchor, drawing_ns)

                        # 方法1: 単一の図形要素を探す
                        single_shapes = anchor.findall('./xdr:sp', drawing_ns)
                        for shape in single_shapes:
                            self._process_shape_from_xml(
                                shape, drawing_ns, anchor_type, anchor_position,
                                shapes_md, shapes_info
                            )

                        # 方法2: グループ化された図形を探す
                        group_shapes = anchor.findall('./xdr:grpSp', drawing_ns)
                        for group in group_shapes:
                            # グループ内のすべての図形を取得
                            group_shapes_list = group.findall('.//xdr:sp', drawing_ns)
                            for shape in group_shapes_list:
                                self._process_shape_from_xml(
                                    shape, drawing_ns, anchor_type, anchor_position,
                                    shapes_md, shapes_info, is_grouped=True
                                )

        except Exception as e:
            print(f"ZIP解析による図形抽出エラー: {str(e)}")
            if self.config.get('verbose_logging', False):
                import traceback
                traceback.print_exc()

        return shapes_md, shapes_info

    def _process_shape_from_xml(self, shape, drawing_ns: dict, anchor_type: str,
                                anchor_position: str, shapes_md: list, shapes_info: list,
                                is_grouped: bool = False):
        """
        XML要素から図形データを抽出してリストに追加

        Args:
            shape: 図形のXML要素
            drawing_ns: XML名前空間の辞書
            anchor_type: アンカータイプ
            anchor_position: アンカーの位置情報
            shapes_md: Markdownリスト（出力先）
            shapes_info: 図形情報リスト（出力先）
            is_grouped: グループ化された図形かどうか
        """
        self.shape_counter += 1

        # 図形名を取得
        shape_name = f"Shape {self.shape_counter}"
        nv_sp_pr = shape.find('.//xdr:nvSpPr', drawing_ns)
        if nv_sp_pr is not None:
            c_nv_pr = nv_sp_pr.find('.//xdr:cNvPr', drawing_ns)
            if c_nv_pr is not None:
                name_attr = c_nv_pr.get('name')
                if name_attr:
                    shape_name = name_attr

        # テキストを取得
        text_parts = []
        tx_body = shape.find('.//xdr:txBody', drawing_ns)

        if tx_body is not None:
            paragraphs = tx_body.findall('.//a:p', drawing_ns)

            for paragraph in paragraphs:
                para_text = []

                # テキストラン（a:r）を取得
                runs = paragraph.findall('.//a:r', drawing_ns)
                for run in runs:
                    t_elem = run.find('.//a:t', drawing_ns)
                    if t_elem is not None and t_elem.text:
                        para_text.append(t_elem.text)

                if para_text:
                    text_parts.append(''.join(para_text))

        shape_text = '\n'.join(text_parts) if text_parts else None

        # テキストがある場合のみ追加
        if shape_text:
            shape_data = {
                'index': self.shape_counter,
                'name': shape_name,
                'type': 'shape',
                'anchor_type': anchor_type,
                'text': shape_text,
                'is_grouped': is_grouped
            }

            # 位置情報を追加
            if anchor_position:
                shape_data['position'] = anchor_position

            # Markdown形式で出力
            group_indicator = " (グループ化)" if is_grouped else ""
            md_parts = [f"### 📐 {shape_name}{group_indicator}"]
            for line in shape_text.split('\n'):
                if line.strip():
                    md_parts.append(f"> {line}")

            if anchor_position:
                md_parts.append(f"\n**位置情報**: {anchor_position}")

            md_shape = '\n'.join(md_parts)
            shapes_md.append(md_shape)
            shapes_info.append(shape_data)

    def _get_position_from_xml_anchor(self, anchor, ns: dict) -> str:
        """XMLアンカーから位置情報を取得"""
        try:
            # twoCellAnchorの場合
            from_elem = anchor.find('.//xdr:from', ns)
            if from_elem is not None:
                col_elem = from_elem.find('.//xdr:col', ns)
                row_elem = from_elem.find('.//xdr:row', ns)

                if col_elem is not None and row_elem is not None:
                    col = int(col_elem.text) if col_elem.text else 0
                    row = int(row_elem.text) if row_elem.text else 0

                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col + 1)
                    return f"セル {col_letter}{row + 1} 付近"

            return ""
        except Exception:
            return ""

    def _get_anchor_info(self, anchor) -> str:
        """図形の位置情報を取得"""
        try:
            # twoCellAnchorの場合
            if hasattr(anchor, '_from'):
                from_cell = anchor._from
                if hasattr(from_cell, 'col') and hasattr(from_cell, 'row'):
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(from_cell.col + 1)
                    return f"セル {col_letter}{from_cell.row + 1} 付近"
            # 別の方法でアンカー情報を取得
            elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(anchor.col + 1)
                return f"セル {col_letter}{anchor.row + 1} 付近"
            return ""
        except Exception:
            return ""

    def _save_image(self, img, output_path: str):
        """画像を保存"""
        try:
            # openpyxlの画像オブジェクトからPIL Imageに変換
            if hasattr(img, '_data'):
                # 画像データを取得
                image_data = img._data()
                pil_image = Image.open(io.BytesIO(image_data))

                # 最大サイズの制限（設定により）
                max_size = tuple(self.config.get('max_size', [1920, 1080]))
                if pil_image.size[0] > max_size[0] or pil_image.size[1] > max_size[1]:
                    pil_image.thumbnail(max_size, Image.LANCZOS)

                # ファイル形式に応じて保存
                image_format = self.config.get('image_format', 'png').upper()
                if image_format == 'JPG':
                    image_format = 'JPEG'

                pil_image.save(output_path, format=image_format)

        except Exception as e:
            print(f"画像保存エラー: {str(e)}")
            raise

    def _generate_image_description(self, img) -> str:
        """画像の説明を生成（基本的な情報のみ）"""
        description_parts = ["【画像情報】"]

        # 画像名
        if hasattr(img, 'name') and img.name:
            description_parts.append(f"- 名前: {img.name}")

        # 画像サイズ
        if hasattr(img, 'width') and hasattr(img, 'height'):
            description_parts.append(f"- サイズ: {img.width} x {img.height}")

        if len(description_parts) > 1:
            return '\n'.join(description_parts)

        return ""
