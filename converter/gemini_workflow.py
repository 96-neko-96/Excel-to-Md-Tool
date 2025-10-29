"""
Gemini Workflow Manager
ExcelシートをPDF/画像に変換し、Gemini APIで分析する統合ワークフロー
"""

import os
import tempfile
import shutil
from typing import Dict, List, Optional, Callable
import openpyxl
from .sheet_to_image import SheetToImageConverter
from .gemini_analyzer import GeminiAnalyzer


class GeminiWorkflowManager:
    """
    Excel → 画像 → Gemini分析の統合ワークフローを管理するクラス
    """

    def __init__(self, gemini_api_key: str, dpi: int = 150, model_name: str = "gemini-2.5-flash-lite"):
        """
        初期化

        Args:
            gemini_api_key: Gemini APIキー
            dpi: 画像の解像度（デフォルト: 150）
            model_name: 使用するGeminiモデル名（デフォルト: gemini-2.5-flash-lite）
        """
        self.gemini_api_key = gemini_api_key
        self.dpi = dpi
        self.model_name = model_name

        # コンバーターとアナライザーの初期化
        self.image_converter = SheetToImageConverter(dpi=dpi)
        self.gemini_analyzer = GeminiAnalyzer(api_key=gemini_api_key, model_name=model_name)

    def process_excel_file(
        self,
        excel_path: str,
        output_dir: str,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> Dict:
        """
        Excelファイル全体を処理（画像変換 + Gemini分析）

        Args:
            excel_path: Excelファイルのパス
            output_dir: 出力ディレクトリ
            progress_callback: 進捗コールバック関数 (current, total, message)

        Returns:
            処理結果
            {
                "sheets": {
                    "シート名": {
                        "image_path": "画像パス",
                        "analysis": {...},
                        "markdown": "Markdown形式の分析結果"
                    },
                    ...
                },
                "summary": {
                    "total_sheets": 総シート数,
                    "processed_sheets": 処理済みシート数,
                    "failed_sheets": 失敗シート数
                }
            }
        """
        os.makedirs(output_dir, exist_ok=True)

        # 画像出力ディレクトリ
        images_dir = os.path.join(output_dir, "images")
        os.makedirs(images_dir, exist_ok=True)

        # Markdown出力ディレクトリ
        markdown_dir = os.path.join(output_dir, "markdown")
        os.makedirs(markdown_dir, exist_ok=True)

        # Excelファイルを開く
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_names = workbook.sheetnames
        total_sheets = len(sheet_names)

        results = {
            "sheets": {},
            "summary": {
                "total_sheets": total_sheets,
                "processed_sheets": 0,
                "failed_sheets": 0
            }
        }

        # 各シートを処理
        for idx, sheet_name in enumerate(sheet_names, 1):
            if progress_callback:
                progress_callback(idx, total_sheets, f"{sheet_name} を処理中...")

            try:
                sheet = workbook[sheet_name]

                # ステップ1: シートを画像に変換
                image_path = self.image_converter._sheet_to_image(
                    sheet, images_dir, sheet_name
                )

                # ステップ2: Geminiで分析
                analysis_results = self.gemini_analyzer.analyze_full_sheet(image_path)

                # ステップ3: Markdownを生成
                markdown_content = self.gemini_analyzer.generate_markdown_from_analysis(
                    analysis_results, sheet_name
                )

                # Markdownファイルを保存
                markdown_path = os.path.join(
                    markdown_dir,
                    f"{self.image_converter._sanitize_filename(sheet_name)}.md"
                )
                with open(markdown_path, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)

                # 結果を保存
                results["sheets"][sheet_name] = {
                    "image_path": image_path,
                    "analysis": analysis_results,
                    "markdown": markdown_content,
                    "markdown_path": markdown_path
                }

                results["summary"]["processed_sheets"] += 1

            except Exception as e:
                print(f"シート '{sheet_name}' の処理中にエラー: {e}")
                results["sheets"][sheet_name] = {
                    "error": str(e)
                }
                results["summary"]["failed_sheets"] += 1

        workbook.close()

        # 統合Markdownファイルを生成
        combined_markdown_path = os.path.join(output_dir, "combined_analysis.md")
        self._create_combined_markdown(results, combined_markdown_path)

        return results

    def process_single_sheet(
        self,
        excel_path: str,
        sheet_name: str,
        output_dir: str
    ) -> Dict:
        """
        単一シートを処理

        Args:
            excel_path: Excelファイルのパス
            sheet_name: 処理するシート名
            output_dir: 出力ディレクトリ

        Returns:
            処理結果
        """
        os.makedirs(output_dir, exist_ok=True)

        # 画像出力ディレクトリ
        images_dir = os.path.join(output_dir, "images")
        os.makedirs(images_dir, exist_ok=True)

        # Excelファイルを開く
        workbook = openpyxl.load_workbook(excel_path, data_only=True)

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません")

        sheet = workbook[sheet_name]

        try:
            # ステップ1: シートを画像に変換
            image_path = self.image_converter._sheet_to_image(
                sheet, images_dir, sheet_name
            )

            # ステップ2: Geminiで分析
            analysis_results = self.gemini_analyzer.analyze_full_sheet(image_path)

            # ステップ3: Markdownを生成
            markdown_content = self.gemini_analyzer.generate_markdown_from_analysis(
                analysis_results, sheet_name
            )

            result = {
                "image_path": image_path,
                "analysis": analysis_results,
                "markdown": markdown_content
            }

        finally:
            workbook.close()

        return result

    def _create_combined_markdown(self, results: Dict, output_path: str):
        """
        全シートの分析結果を統合したMarkdownファイルを生成

        Args:
            results: process_excel_fileの結果
            output_path: 出力ファイルパス
        """
        lines = []

        # タイトル
        lines.append("# Excel分析レポート（Gemini AI分析）")
        lines.append("")

        # サマリー
        summary = results["summary"]
        lines.append("## サマリー")
        lines.append(f"- 総シート数: {summary['total_sheets']}")
        lines.append(f"- 処理成功: {summary['processed_sheets']}")
        lines.append(f"- 処理失敗: {summary['failed_sheets']}")
        lines.append("")

        # 目次
        lines.append("## 目次")
        for sheet_name in results["sheets"].keys():
            if "error" not in results["sheets"][sheet_name]:
                lines.append(f"- [{sheet_name}](#{self._create_anchor(sheet_name)})")
        lines.append("")
        lines.append("---")
        lines.append("")

        # 各シートの分析結果
        for sheet_name, sheet_data in results["sheets"].items():
            if "error" in sheet_data:
                lines.append(f"## {sheet_name}")
                lines.append(f"**エラー:** {sheet_data['error']}")
                lines.append("")
                continue

            # Markdownコンテンツを追加
            markdown_content = sheet_data.get("markdown", "")
            lines.append(markdown_content)
            lines.append("")

        # ファイルに書き込み
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))

    @staticmethod
    def _create_anchor(text: str) -> str:
        """
        Markdownのアンカーリンク用のIDを生成

        Args:
            text: 元のテキスト

        Returns:
            アンカーID
        """
        # 小文字化、空白をハイフンに
        anchor = text.lower().replace(" ", "-")
        # 特殊文字を除去
        anchor = "".join(c for c in anchor if c.isalnum() or c == "-")
        return anchor
