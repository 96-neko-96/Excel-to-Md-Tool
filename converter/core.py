"""
Excel to Markdown Converter Core
"""

import os
import warnings
import openpyxl
from typing import Dict, List, Any
from datetime import datetime

from .sheet_parser import SheetParser
from .markdown_generator import MarkdownGenerator
from .metadata_generator import MetadataGenerator
from .image_parser import ImageParser


class ExcelToMarkdownConverter:
    """Excel to Markdown 変換エンジン"""

    def __init__(self, **config):
        """
        Args:
            chunk_size: RAG用チャンクサイズ
            create_toc: 目次生成の有効化
            extract_images: 画像抽出の有効化
            generate_summary: 表要約の有効化
            enable_ai_features: AI機能の有効化 (Phase 3)
            ai_table_summary: 表の自然言語要約 (Phase 3)
            ai_image_description: 画像の説明自動生成 (Phase 3)
            ai_generate_qa: よくあるQA生成 (Phase 3)
            gemini_api_key: Gemini APIキー (Phase 3)
            gemini_model: Geminiモデル名 (Phase 3)
        """
        self.config = {
            'chunk_size': config.get('chunk_size', 800),
            'create_toc': config.get('create_toc', True),
            'extract_images': config.get('extract_images', True),
            'generate_summary': config.get('generate_summary', False),
            'show_formulas': config.get('show_formulas', True),
            # Phase 3: AI機能
            'enable_ai_features': config.get('enable_ai_features', False),
            'ai_table_summary': config.get('ai_table_summary', False),
            'ai_image_description': config.get('ai_image_description', False),
            'ai_generate_qa': config.get('ai_generate_qa', False),
            'gemini_api_key': config.get('gemini_api_key', ''),
            'gemini_model': config.get('gemini_model', 'gemini-2.5-flash-lite'),
        }
        # その他の設定をマージ
        self.config.update(config)

        self.workbook = None
        self.workbook_with_values = None  # 実数値用
        self.sheets_data = []
        self.cross_references = []

        # パーサーの初期化
        self.sheet_parser = SheetParser(self.config, gemini_analyzer=None)  # 後で設定
        self.markdown_generator = MarkdownGenerator(self.config)
        self.metadata_generator = MetadataGenerator(self.config)
        self.image_parser = ImageParser(self.config)

        # Phase 3: Gemini Analyzerの初期化（AI機能が有効な場合のみ）
        self.gemini_analyzer = None
        if self.config.get('enable_ai_features') and self.config.get('gemini_api_key'):
            try:
                from .gemini_analyzer import GeminiAnalyzer
                self.gemini_analyzer = GeminiAnalyzer(
                    api_key=self.config['gemini_api_key'],
                    model_name=self.config['gemini_model']
                )
                # SheetParserにGeminiAnalyzerを設定
                self.sheet_parser.set_gemini_analyzer(self.gemini_analyzer)
            except Exception as e:
                print(f"Gemini Analyzerの初期化エラー: {e}")
                self.gemini_analyzer = None

    def convert(self, input_path: str, output_path: str) -> Dict[str, Any]:
        """
        Excelファイルを変換

        Args:
            input_path: 入力Excelファイルパス
            output_path: 出力Markdownファイルパス

        Returns:
            変換結果の統計情報
        """
        # 1. Excel読み込み（数式用と実数値用の2つ）
        self.workbook, self.workbook_with_values = self._load_excel(input_path)

        # 2. 各シートを変換
        for idx, sheet in enumerate(self.workbook.worksheets):
            # 非表示シートはスキップ（設定による）
            if sheet.sheet_state == 'hidden' and not self.config.get('include_hidden', False):
                continue

            # 対応する実数値シートを取得
            sheet_with_values = self.workbook_with_values.worksheets[idx] if self.workbook_with_values else None

            sheet_data = self._convert_sheet(sheet, sheet_with_values)
            self.sheets_data.append(sheet_data)

        # 3. シート間参照解析
        self.cross_references = self._analyze_references()

        # 4. 統合Markdown生成
        md_content = self._merge_sheets()

        # 5. ファイル出力
        self._write_output(output_path, md_content)

        # 6. メタデータ生成
        metadata = self._generate_metadata(input_path, output_path)

        return {
            'sheets_count': len(self.sheets_data),
            'tables_count': sum(s['tables_count'] for s in self.sheets_data),
            'images_count': sum(s['images_count'] for s in self.sheets_data),
            'shapes_count': sum(s['shapes_count'] for s in self.sheets_data),
            'estimated_chunks': self._estimate_chunks(md_content),
            'metadata': metadata,
            'output_file': output_path
        }

    def _load_excel(self, path: str) -> tuple:
        """Excelファイルを読み込む（数式用と実数値用の2つ）"""
        try:
            # openpyxlの警告を抑制（データ検証などの非サポート機能）
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

                # 数式情報を保持して読み込み
                workbook_formulas = openpyxl.load_workbook(path, data_only=False)

                # 実数値のみを読み込み
                workbook_values = openpyxl.load_workbook(path, data_only=True)

            return workbook_formulas, workbook_values
        except Exception as e:
            raise ValueError(f"Excelファイルの読み込みに失敗しました: {str(e)}")

    def _convert_sheet(self, sheet, sheet_with_values=None) -> Dict[str, Any]:
        """シートを変換"""
        return self.sheet_parser.parse_sheet(sheet, sheet_with_values)

    def _analyze_references(self) -> List[Dict[str, Any]]:
        """シート間参照を解析"""
        references = []

        if not self.workbook:
            return references

        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # 数式セル
                        formula = str(cell.value)

                        # 他シート参照の検出
                        if '!' in formula:
                            ref_info = self._parse_cross_sheet_reference(formula, sheet.title, cell.coordinate)
                            if ref_info:
                                references.append(ref_info)

        return references

    def _parse_cross_sheet_reference(self, formula: str, from_sheet: str, from_cell: str) -> Dict[str, Any]:
        """数式から他シート参照を解析"""
        import re

        # 例: "=SUM(Sheet1!A1:A10)" または "='Sheet Name'!A1"
        pattern = r"(['\"]?)([^'\"!]+)\1!([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)"
        matches = re.findall(pattern, formula)

        if matches:
            # 最初の参照のみを取得
            _, sheet_name, cell_range = matches[0]
            return {
                'from_sheet': from_sheet,
                'from_cell': from_cell,
                'to_sheet': sheet_name,
                'to_cell': cell_range,
                'formula': formula
            }

        return None

    def _merge_sheets(self) -> str:
        """複数シートを統合"""
        return self.markdown_generator.merge_sheets(
            self.sheets_data,
            self.cross_references,
            self.workbook
        )

    def _write_output(self, path: str, content: str):
        """Markdownファイルを出力"""
        try:
            output_dir = os.path.dirname(path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)
        except Exception as e:
            raise IOError(f"ファイルの書き込みに失敗しました: {str(e)}")

    def _generate_metadata(self, input_path: str, output_path: str) -> Dict[str, Any]:
        """メタデータを生成"""
        return self.metadata_generator.generate(
            self.workbook,
            self.sheets_data,
            self.cross_references,
            input_path,
            output_path
        )

    def _estimate_chunks(self, content: str) -> int:
        """推奨チャンク数を推定"""
        # 簡易的な推定（1文字 ≒ 0.3トークン程度）
        estimated_tokens = len(content) * 0.3
        chunk_size = self.config['chunk_size']

        return max(1, int(estimated_tokens // chunk_size) + 1)
