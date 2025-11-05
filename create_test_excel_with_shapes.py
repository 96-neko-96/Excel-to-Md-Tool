"""
図形を含むExcelファイルを作成するスクリプト
注意: Excelアプリケーションで開いて図形を追加する必要があります
"""
import openpyxl

def create_excel_with_textbox():
    """
    テキストボックスを含むExcelファイルを作成

    注意: openpyxlでは図形の作成が完全にサポートされていないため、
    このファイルをExcelで開いて手動で図形を追加することをお勧めします。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テストシート"

    # データを追加
    ws['A1'] = "商品名"
    ws['B1'] = "価格"
    ws['C1'] = "説明"

    ws['A2'] = "商品A"
    ws['B2'] = 1000
    ws['C2'] = "人気商品"

    ws['A3'] = "商品B"
    ws['B3'] = 2000
    ws['C3'] = "新商品"

    ws['A5'] = "注意事項"
    ws['A6'] = "図形を追加するには、Excelで開いて「挿入」→「図形」→「テキストボックス」を選択してください"

    # ファイルを保存
    output_path = "/tmp/sample_for_shapes.xlsx"
    wb.save(output_path)

    print(f"✓ サンプルExcelファイルを作成しました: {output_path}")
    print()
    print("次のステップ:")
    print("1. Excelでこのファイルを開く")
    print("2. 「挿入」→「図形」→「テキストボックス」を選択")
    print("3. シート上にテキストボックスを描画")
    print("4. テキストボックスに「これはテスト用のテキストボックスです」などのテキストを入力")
    print("5. ファイルを保存")
    print("6. 保存したファイルを使ってテストを実行")
    print()
    print(f"テスト実行コマンド:")
    print(f"  python test_shape_extraction.py {output_path}")

    return output_path

if __name__ == "__main__":
    create_excel_with_textbox()
